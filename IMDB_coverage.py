import shutil
import warnings
import sys
import time
from openpyxl import Workbook, load_workbook
import pickle
from sklearn.preprocessing import LabelEncoder, OneHotEncoder
import tensorflow as tf
from keras.models import load_model
from keras import backend as K
import numpy as np
from coverage import Coverage
import os
import pymannkendall as mk
import pandas as pd
from tensorflow.keras.preprocessing.sequence import pad_sequences
from sklearn.utils import shuffle

warnings.filterwarnings("ignore")
config = tf.compat.v1.ConfigProto(gpu_options=tf.compat.v1.GPUOptions(allow_growth=True))
sess = tf.compat.v1.Session(config=config)


def encode_text(tokenizer, lines, length):
    # integer encode
    encoded = tokenizer.texts_to_sequences(lines)
    # pad encoded sequences
    padded = pad_sequences(encoded, maxlen=length, padding="post")
    return padded


def load_data():
    train_df = pd.read_csv(
        "./IMDB/train.csv",
    )

    test_df = pd.read_csv(
        "./IMDB/test.csv",
    )

    train_df = train_df.sample(frac=1).reset_index(drop=True)
    test_df = test_df.sample(frac=1).reset_index(drop=True)

    train_x = train_df.text.tolist()
    train_y = train_df.label

    test_x = test_df.text.tolist()
    test_y = test_df.label

    max_len = 200
    # tokenizer = create_tokenizer(train_x)
    with open("./tokenizer/200in_IMDB.pickle", "rb") as handle:
        tokenizer = pickle.load(handle)

    # calculate vocabulary size
    vocab_size = len(tokenizer.word_index) + 1
    word_index = tokenizer.word_index
    train_x = encode_text(tokenizer, train_x, max_len)

    test_x = encode_text(tokenizer, test_x, max_len)

    le = LabelEncoder()
    train_y = le.fit_transform(train_y).reshape(-1, 1)

    test_y = le.transform(test_y).reshape(-1, 1)

    return train_x, train_y, test_x, test_y


if __name__ == '__main__':

    netname = "CNNLSTM"
    dataname = "IMDB"

    epoch = 10
    num = 25000
    seed = 42
    T = 0
    flag = 0

    x_train, y_train, x_test, y_test = load_data()

    data_shape = [x_train.shape]
    print(data_shape)

    # x_train, y_train = shuffle(x_train, y_train, random_state=seed)

    model_path = './Model/' + netname + '_' + dataname + '/'
    dirs = os.listdir(model_path)
    model = load_model(model_path + dirs[0])

    model_layer = len(model.layers) - 1
    data_shape = x_train.shape
    # print(data_shape)
    # print(model_layer)

    l = []
    model.summary()
    for i in range(1, model_layer):
        if model.layers[i].output.shape[1] != None:
            l.append(i)

    # 记录每次迭代的coverage list
    trend_train_nc1 = []
    trend_train_nc2 = []
    trend_train_nc3 = []
    trend_train_nc4 = []
    trend_train_nc5 = []
    trend_train_tknc = []
    trend_train_kmnc = []
    trend_train_tknp = []
    trend_train_nbc = []
    trend_train_snac = []
    trend_test_nc1 = []
    trend_test_nc2 = []
    trend_test_nc3 = []
    trend_test_nc4 = []
    trend_test_nc5 = []
    trend_test_tknc = []
    trend_test_kmnc = []
    trend_test_tknp = []
    trend_test_nbc = []
    trend_test_snac = []

    for i in dirs:
        print(flag)
        flag = flag + 1
        print(i)
        print(time.asctime())
        model = load_model(model_path + i)
        # 计算train_data上的coverage
        # coverage = Coverage(model, x_train[0: num], x_train[0: num])
        # nc1, activate_num1, total_num1 = coverage.NC(l, threshold=0.1)
        # nc2, activate_num2, total_num2 = coverage.NC(l, threshold=0.3)
        # nc3, activate_num3, total_num3 = coverage.NC(l, threshold=0.5)
        # nc4, activate_num4, total_num4 = coverage.NC(l, threshold=0.7)
        # nc5, activate_num5, total_num5 = coverage.NC(l, threshold=0.9)
        # tknc, pattern_num, total_num6 = coverage.TKNC(l)
        # tknp = coverage.TKNP(l)
        # kmnc, nbc, snac, covered_num, l_covered_num, u_covered_num, neuron_num = coverage.KMNC(l)
        #
        # trend_train_nc1.append(nc1)
        # trend_train_nc2.append(nc2)
        # trend_train_nc3.append(nc3)
        # trend_train_nc4.append(nc4)
        # trend_train_nc5.append(nc5)
        # trend_train_tknc.append(tknc)
        # trend_train_kmnc.append(kmnc)
        # trend_train_tknp.append(tknp)
        # trend_train_nbc.append(nbc)
        # trend_train_snac.append(snac)

        # with open("./Coverage/coverage_result_of_train_data_" + netname + "_" + dataname + ".txt", "a") as f:
        #     f.write("\n------------------------------------------------------------------------------\n")
        #     f.write('the result of {} {} epoch_{} is: \n'.format(netname, dataname, i))
        #     f.write('NC(0.1): {}  activate_num: {}  total_num: {} \n'.format(nc1, activate_num1, total_num1))
        #     f.write('NC(0.3): {}  activate_num: {}  total_num: {} \n'.format(nc2, activate_num2, total_num2))
        #     f.write('NC(0.5): {}  activate_num: {}  total_num: {} \n'.format(nc3, activate_num3, total_num3))
        #     f.write('NC(0.7): {}  activate_num: {}  total_num: {} \n'.format(nc4, activate_num4, total_num4))
        #     f.write('NC(0.9): {}  activate_num: {}  total_num: {} \n'.format(nc5, activate_num5, total_num5))
        #     f.write('TKNC: {}  pattern_num: {}  total_num: {} \n'.format(tknc, pattern_num, total_num6))
        #     f.write('TKNP: {} \n'.format(tknp))
        #     f.write('KMNC: {}  covered_num: {}  total_num: {} \n'.format(kmnc, covered_num, neuron_num))
        #     f.write('NBC: {}  l_covered_num: {}  u_covered_num: {} \n'.format(nbc, l_covered_num, u_covered_num))
        #     f.write('SNAC: {} \n'.format(snac))
        #
        # K.clear_session()

        # 计算test_data上的coverage
        coverage = Coverage(model, x_train[0: num], x_test[0: num])
        nc1, activate_num1, total_num1 = coverage.NC(l, threshold=0.1)
        nc2, activate_num2, total_num2 = coverage.NC(l, threshold=0.3)
        nc3, activate_num3, total_num3 = coverage.NC(l, threshold=0.5)
        nc4, activate_num4, total_num4 = coverage.NC(l, threshold=0.7)
        nc5, activate_num5, total_num5 = coverage.NC(l, threshold=0.9)
        tknc, pattern_num, total_num6 = coverage.TKNC(l)
        tknp = coverage.TKNP(l)
        kmnc, nbc, snac, covered_num, l_covered_num, u_covered_num, neuron_num = coverage.KMNC(l)

        trend_test_nc1.append(nc1)
        trend_test_nc2.append(nc2)
        trend_test_nc3.append(nc3)
        trend_test_nc4.append(nc4)
        trend_test_nc5.append(nc5)
        trend_test_tknc.append(tknc)
        trend_test_kmnc.append(kmnc)
        trend_test_tknp.append(tknp)
        trend_test_nbc.append(nbc)
        trend_test_snac.append(snac)

        with open("./Coverage/coverage_result_of_test_data_" + netname + "_" + dataname + ".txt", "a") as f:
            f.write("\n------------------------------------------------------------------------------\n")
            f.write('the result of {} {} epoch{} is: \n'.format(netname, dataname, flag))
            f.write('NC(0.1): {}  activate_num: {}  total_num: {} \n'.format(nc1, activate_num1, total_num1))
            f.write('NC(0.3): {}  activate_num: {}  total_num: {} \n'.format(nc2, activate_num2, total_num2))
            f.write('NC(0.5): {}  activate_num: {}  total_num: {} \n'.format(nc3, activate_num3, total_num3))
            f.write('NC(0.7): {}  activate_num: {}  total_num: {} \n'.format(nc4, activate_num4, total_num4))
            f.write('NC(0.9): {}  activate_num: {}  total_num: {} \n'.format(nc5, activate_num5, total_num5))
            f.write('TKNC: {}  pattern_num: {}  total_num: {} \n'.format(tknc, pattern_num, total_num6))
            f.write('TKNP: {} \n'.format(tknp))
            f.write('KMNC: {}  covered_num: {}  total_num: {} \n'.format(kmnc, covered_num, neuron_num))
            f.write('NBC: {}  l_covered_num: {}  u_covered_num: {} \n'.format(nbc, l_covered_num, u_covered_num))
            f.write('SNAC: {} \n'.format(snac))
        K.clear_session()

    # train_trend = [['NC0.1'] + trend_train_nc1,
    #                ['NC0.3'] + trend_train_nc2,
    #                ['NC0.5'] + trend_train_nc3,
    #                ['NC0.7'] + trend_train_nc4,
    #                ['NC0.9'] + trend_train_nc5,
    #                ['TKNC'] + trend_train_tknc,
    #                ['TKNP'] + trend_train_tknp,
    #                ['KMNC'] + trend_train_kmnc,
    #                ['NBC'] + trend_train_nbc,
    #                ['SNAC'] + trend_train_snac]
    #
    if not os.path.isfile('./Coverage/{}_{}.xlsx'.format(netname, dataname)):
        wb = Workbook()
    else:
        wb = load_workbook('./Coverage/{}_{}.xlsx'.format(netname, dataname))
    # ws = wb.create_sheet('train {}'.format(T))
    # for row in train_trend:
    #     ws.append(row)
    # wb.save('./Coverage/{}_{}.xlsx'.format(netname, dataname))

    test_trend = [['NC0.1'] + trend_test_nc1,
                  ['NC0.3'] + trend_test_nc2,
                  ['NC0.5'] + trend_test_nc3,
                  ['NC0.7'] + trend_test_nc4,
                  ['NC0.9'] + trend_test_nc5,
                  ['TKNC'] + trend_test_tknc,
                  ['TKNP'] + trend_test_tknp,
                  ['KMNC'] + trend_test_kmnc,
                  ['NBC'] + trend_test_nbc,
                  ['SNAC'] + trend_test_snac]

    ws = wb.create_sheet('test {}'.format(T))
    for row in test_trend:
        ws.append(row)
    wb.save('./Coverage/{}_{}.xlsx'.format(netname, dataname))

    # print(trend_test_nc1,trend_train_snac)
    """
    trend_result_train_nc1 = mk.original_test(trend_train_nc1)
    trend_result_train_nc2 = mk.original_test(trend_train_nc2)
    trend_result_train_nc3 = mk.original_test(trend_train_nc3)
    trend_result_train_nc4 = mk.original_test(trend_train_nc4)
    trend_result_train_nc5 = mk.original_test(trend_train_nc5)
    trend_result_train_tknc = mk.original_test(trend_train_tknc)
    trend_result_train_tknp = mk.original_test(trend_train_tknp)
    trend_result_train_kmnc = mk.original_test(trend_train_kmnc)
    trend_result_train_nbc = mk.original_test(trend_train_nbc)
    trend_result_train_snac = mk.original_test(trend_train_snac)
    with open("coverage_trend_of_train_data_mnist.txt", "a") as f:
        f.write("\n------------------------------------------------------------------------------\n")
        f.write("data number = {} \n".format(num))
        f.write('the result of {} {} experiment{} is: \n'.format(dataset, model_name, R))
        f.write('NC(0.1): {}  \n\n'.format(trend_result_train_nc1))
        f.write('NC(0.3): {}  \n\n'.format(trend_result_train_nc2))
        f.write('NC(0.5): {}  \n\n'.format(trend_result_train_nc3))
        f.write('NC(0.7): {}  \n\n'.format(trend_result_train_nc4))
        f.write('NC(0.9): {}  \n\n'.format(trend_result_train_nc5))
        f.write('TKNC: {}  \n\n'.format(trend_result_train_tknc))
        f.write('TKNP: {}  \n\n'.format(trend_result_train_tknp))
        f.write('KMNC: {}  \n\n'.format(trend_result_train_kmnc))
        f.write('NBC: {}   \n\n'.format(trend_result_train_nbc))
        f.write('SNAC: {}  \n\n'.format(trend_result_train_snac))

    trend_result_test_nc1 = mk.original_test(trend_test_nc1)
    trend_result_test_nc2 = mk.original_test(trend_test_nc2)
    trend_result_test_nc3 = mk.original_test(trend_test_nc3)
    trend_result_test_nc4 = mk.original_test(trend_test_nc4)
    trend_result_test_nc5 = mk.original_test(trend_test_nc5)
    trend_result_test_tknc = mk.original_test(trend_test_tknc)
    trend_result_test_tknp = mk.original_test(trend_test_tknp)
    trend_result_test_kmnc = mk.original_test(trend_test_kmnc)
    trend_result_test_nbc = mk.original_test(trend_test_nbc)
    trend_result_test_snac = mk.original_test(trend_test_snac)
    with open("coverage_trend_of_test_data.txt", "a") as f:
        f.write("\n------------------------------------------------------------------------------\n")
        f.write("data number = {} \n".format(num))
        f.write('the result of {} {} experiment{} is: \n'.format(dataset, model_name, R))
        f.write('NC(0.1): {}  \n\n'.format(trend_result_test_nc1))
        f.write('NC(0.3): {}  \n\n'.format(trend_result_test_nc2))
        f.write('NC(0.5): {}  \n\n'.format(trend_result_test_nc3))
        f.write('NC(0.7): {}  \n\n'.format(trend_result_test_nc4))
        f.write('NC(0.9): {}  \n\n'.format(trend_result_test_nc5))
        f.write('TKNC: {}  \n\n'.format(trend_result_test_tknc))
        f.write('TKNP: {}  \n\n'.format(trend_result_test_tknp))
        f.write('KMNC: {}  \n\n'.format(trend_result_test_kmnc))
        f.write('NBC: {}   \n\n'.format(trend_result_test_nbc))
        f.write('SNAC: {}  \n\n'.format(trend_result_test_snac))
    """
